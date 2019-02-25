from openpyxl import load_workbook
import sys

NUM_COLUMNS = 50
NUM_ROWS = 32

COLOR_GREEN = 'FF5DBF29'
COLOR_BLUE = 'FF00B0F0'
COLOR_BLACK = '1'

CODE_GREEN = '0'
CODE_BLUE = '1'
CODE_BLACK = '2'

wb = load_workbook(sys.argv[1] + '.xlsx')
ws = wb.active

with open(sys.argv[1] + '.txt', 'w') as f:
    for row in range (1, NUM_ROWS+1):
        for column in range(0,NUM_COLUMNS):
            cell = ws[row][column]
            cell_color = str(cell.fill.fgColor.index)

            if cell_color == COLOR_GREEN:
                f.write(CODE_GREEN)
                print (' ', end='')
            elif cell_color == COLOR_BLUE:
                f.write(CODE_BLUE)
                print('░', end='')
            elif cell_color == COLOR_BLACK:
                f.write(CODE_BLACK)
                print('█', end='')
            else:
                raise 'Color no definido'
        f.write('\n')
        print()